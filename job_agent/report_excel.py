"""Microsoft Excel (.xlsx) report generation with openpyxl."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any
from collections.abc import Callable

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from . import config, profile
from .models import Job, RunStats
from .pipeline import RunResult
from .utils import local_timezone, safe_url


HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

HOT_FILL = PatternFill("solid", fgColor="FFF2CC")
FRESH_FILL = PatternFill("solid", fgColor="C6EFCE")
NEW_FILL = PatternFill("solid", fgColor="DDEBF7")
SEEN_FILL = PatternFill("solid", fgColor="F2F2F2")
PROSPECT_FILL = PatternFill("solid", fgColor="FCE4D6")

NEW_FONT = Font(name="Calibri", size=11, bold=True, color="1F3864")
LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")
BASE_FONT = Font(name="Calibri", size=11)

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

DATE_FORMAT = "yyyy-mm-dd"
CURRENCY_FORMATS = {
    "GBP": '£#,##0', "USD": '$#,##0', "EUR": '€#,##0',
    "CAD": '"C$"#,##0', "AUD": '"A$"#,##0', "CHF": '"CHF "#,##0',
    "PLN": '#,##0" zł"', "INR": '"₹"#,##0', "SEK": '#,##0" kr"',
}
DEFAULT_MONEY_FORMAT = "#,##0"

MAX_WIDTH = 52
MIN_WIDTH = 10


def _join(values: list[str]) -> str:
    return " • ".join(v for v in values if v)


ColumnSpec = tuple[str, Callable[[Job], Any], str]

def _contact(job: Job) -> str:
    """Named person and role, when the advert or the site published one."""
    if job.best_contact_name and job.contact_role:
        return f"{job.best_contact_name} — {job.contact_role}"
    return job.best_contact_name or job.contact_role or ""


COLUMNS: list[ColumnSpec] = [
    ("Shortlist %",    lambda j: j.match_score,                            "score"),
    ("CV Fit %",       lambda j: j.cv_fit_score,                           "score"),
    ("Job Title",      lambda j: j.title,                                  "wrap"),
    ("Company",        lambda j: j.company,                                "wrap"),
    ("Location",       lambda j: j.location,                               "text"),
    ("Arrangement",    lambda j: j.remote_status,                          "text"),
    ("Employment",     lambda j: j.employment_type,                        "short"),
    ("Salary",         lambda j: pay_summary(j),                           "text"),
    ("Posted",         lambda j: j.posted_date or j.job_age_label,         "date"),
    ("Potential gaps", lambda j: j.potential_gaps or _join(j.concerns),    "wrap"),
    ("Email",          lambda j: j.public_email,                           "email"),
    ("Contact",        lambda j: _contact(j),                              "text"),
    ("Website",        lambda j: j.company_website or j.careers_page,      "url"),
    ("Job Link",       lambda j: j.application_url or j.original_job_url,  "url"),
    ("Why this rank",  lambda j: j.chance_explained,                       "wrap"),
]

PROSPECT_COLUMNS: list[ColumnSpec] = COLUMNS + [
    ("Why unconfirmed", lambda j: j.rejection_reason or _join(j.concerns), "wrap"),
]

COL_INDEX = {name: i + 1 for i, (name, _, _) in enumerate(COLUMNS)}

WIDTH_HINTS = {
    "wrap": 34, "url": 34, "text": 18, "short": 16,
    "date": 13, "score": 9, "money": 14, "email": 26, "level": 17, "usd": 16,
}
MAX_WIDTH_BY_KIND = {"wrap": 46, "url": 40, "email": 32, "text": 26, "short": 22,
                     "date": 13, "score": 10, "money": 16, "level": 18, "usd": 18}

LEVEL_STYLES = {
    config.LEVEL_BEGINNER: (PatternFill("solid", fgColor="E2EFDA"), Font(name="Calibri", size=11, bold=True, color="375623")),
    config.LEVEL_MEDIUM:   (PatternFill("solid", fgColor="DDEBF7"), Font(name="Calibri", size=11, bold=True, color="1F4E79")),
    config.LEVEL_SENIOR:   (PatternFill("solid", fgColor="E4DFEC"), Font(name="Calibri", size=11, bold=True, color="4A2E7A")),
}


def _style_header(ws: Worksheet, headers: list[str], row: int = 1) -> None:
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def _hyperlink(cell, url: str, display: str | None = None) -> None:
    safe = safe_url(url)
    if not safe:
        cell.value = display or url or None
        return
    url = safe
    text = display or url
    if len(text) > 60:
        text = text[:57] + "…"
    cell.value = text
    cell.hyperlink = url
    cell.font = LINK_FONT


def _autosize(ws: Worksheet, headers: list[str], kinds: list[str],
              samples: list[list[Any]]) -> None:
    for index, header in enumerate(headers):
        kind = kinds[index] if index < len(kinds) else "text"
        longest = len(header) + 3
        for row in samples:
            value = row[index] if index < len(row) else ""
            if value is None:
                continue
            if isinstance(value, (date, datetime)):
                length = 10
            elif isinstance(value, (int, float)):
                length = len(f"{value:,.0f}") + 2
            else:
                text = str(value)
                length = max((len(part) for part in text.split(" • ")), default=0)
                length = min(length, MAX_WIDTH_BY_KIND.get(kind, MAX_WIDTH))
            longest = max(longest, length)
        width = max(MIN_WIDTH, min(longest + 2, MAX_WIDTH_BY_KIND.get(kind, MAX_WIDTH)))
        ws.column_dimensions[get_column_letter(index + 1)].width = width


def _add_table(ws: Worksheet, name: str, last_row: int, last_col: int) -> None:
    """Attach a real Excel table (which also provides the filter dropdowns)."""
    if last_row < 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}1"
        return
    ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _safe_table_name(sheet_title: str) -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in sheet_title)
    if not cleaned or not cleaned[0].isalpha():
        cleaned = "T_" + cleaned
    return f"{cleaned}_tbl"


def _dedupe(jobs: list[Job]) -> list[Job]:
    seen: set[str] = set()
    out: list[Job] = []
    for job in jobs:
        key = job.fingerprint or f"{job.company}|{job.title}".lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(job)
    return out


def _sorted(jobs: list[Job]) -> list[Job]:
    return sorted(
        jobs,
        key=lambda j: (j.match_score, j.posted_at.timestamp() if j.posted_at else 0),
        reverse=True,
    )


def _cell_url(cell, value, job, is_hot) -> None:
    _hyperlink(cell, value or "")
    cell.alignment = TOP


def _cell_email(cell, value, job, is_hot) -> None:
    if value:
        _hyperlink(cell, f"mailto:{value}", value)
    cell.alignment = TOP


def _cell_date(cell, value, job, is_hot) -> None:
    cell.value = value
    cell.number_format = DATE_FORMAT
    cell.alignment = CENTER


def _cell_money(cell, value, job, is_hot) -> None:
    cell.value = value
    cell.number_format = CURRENCY_FORMATS.get(
        (job.salary_currency or "").upper(), DEFAULT_MONEY_FORMAT)
    cell.alignment = TOP


def _cell_score(cell, value, job, is_hot) -> None:
    cell.value = value
    cell.alignment = CENTER
    cell.font = Font(name="Calibri", size=11, bold=is_hot)


def _cell_usd(cell, value, job, is_hot) -> None:
    cell.value = value
    cell.number_format = '"$"#,##0'
    cell.alignment = TOP
    if value and value >= config.SALARY_FLOOR_USD:
        cell.font = Font(name="Calibri", size=11, bold=True, color="006100")


def _cell_centred(cell, value, job, is_hot) -> None:
    cell.value = value
    cell.alignment = CENTER


#: Leading characters a spreadsheet reads as the start of a formula.
FORMULA_LEADERS = ("=", "+", "-", "@")


def defuse_formulas(wb) -> None:
    """Keep scraped text out of Excel's formula parser.

    Employer names, titles and locations are third-party text. openpyxl reads a
    leading "=" as a formula, so an advert could put one into the reader's
    spreadsheet — `=cmd|'/c calc'!A1` is the usual demonstration. Forcing the
    type back to text keeps every character and drops the behaviour.

    Swept over the finished workbook rather than at each write, so a new sheet
    or column cannot quietly opt out of it.
    """
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(FORMULA_LEADERS):
                    cell.data_type = "s"


def _cell_wrapped(cell, value, job, is_hot) -> None:
    cell.value = value
    cell.alignment = WRAP


def _cell_plain(cell, value, job, is_hot) -> None:
    cell.value = value
    cell.alignment = TOP


# One signature for every writer, so the table reads as a table. Most ignore
# the last two arguments.
CELL_WRITERS = {
    "url": _cell_url,
    "email": _cell_email,
    "date": _cell_date,
    "money": _cell_money,
    "score": _cell_score,
    "usd": _cell_usd,
    "level": _cell_centred,
    "wrap": _cell_wrapped,
}


def _fill_for(job: Job, is_hot: bool):
    """The row tint that says what kind of lead this is."""
    if is_hot:
        return HOT_FILL
    if job.is_prospect:
        return PROSPECT_FILL
    if not job.is_new:
        return SEEN_FILL
    return None


def _write_job_sheet(wb: Workbook, sheet_title: str, jobs: list[Job],
                     subtitle: str = "",
                     columns: list[ColumnSpec] | None = None) -> Worksheet:
    ws = wb.create_sheet(sheet_title)
    columns = columns or COLUMNS
    headers = [name for name, _, _ in columns]
    kinds = [kind for _, _, kind in columns]

    _style_header(ws, headers)

    jobs = _sorted(_dedupe(jobs))
    raw_rows: list[list[Any]] = []

    for row_index, job in enumerate(jobs, start=2):
        is_hot = job.match_score >= config.HOT_LEAD_SCORE
        is_fresh = job.job_age_days is not None and job.job_age_days <= 1
        row_values: list[Any] = []

        for col_index, (header, getter, kind) in enumerate(columns, start=1):
            value = getter(job)
            cell = ws.cell(row=row_index, column=col_index)
            row_values.append(value)

            CELL_WRITERS.get(kind, _cell_plain)(cell, value, job, is_hot)

            cell.border = BORDER

            fill = _fill_for(job, is_hot)
            if fill is not None:
                cell.fill = fill

            if header == "Job Title":
                if job.is_new:
                    cell.font = NEW_FONT
                if is_fresh:
                    cell.fill = FRESH_FILL

        raw_rows.append(row_values)

    last_row = len(jobs) + 1
    last_col = len(columns)

    if not jobs:
        note = ws.cell(row=2, column=1,
                       value=subtitle or "No opportunities matched this category in this run.")
        note.font = Font(name="Calibri", size=11, italic=True, color="808080")

    _autosize(ws, headers, kinds, raw_rows)
    ws.freeze_panes = "A2"
    _add_table(ws, _safe_table_name(sheet_title), last_row, last_col)

    if jobs:
        match_col = get_column_letter(COL_INDEX["Shortlist %"])
        ws.conditional_formatting.add(
            f"{match_col}2:{match_col}{last_row}",
            CellIsRule(operator="greaterThanOrEqual", formula=[str(config.HOT_LEAD_SCORE)],
                       fill=PatternFill("solid", fgColor="FFD966"),
                       font=Font(bold=True, color="7F6000")),
        )

    ws.sheet_view.showGridLines = False
    return ws


COMPANY_COLUMNS: list[tuple[str, str, str]] = [
    ("Company", "company", "text"),
    ("Match %", "best_match_score", "score"),
    ("Roles Advertised", "roles", "wrap"),
    ("Email", "public_email", "email"),
    ("Website", "company_website", "url"),
    ("Careers Page", "careers_page", "url"),
]


def pay_summary(job: Job) -> str:
    """Human-readable pay, e.g. '£75,000 - £95,000' or '$500/day'."""
    symbol = {"GBP": "£", "USD": "$", "EUR": "€"}.get(
        (job.salary_currency or "").upper(), (job.salary_currency or "") + " ")

    def band(low, high, suffix=""):
        if low and high and low != high:
            return f"{symbol}{low:,.0f} - {symbol}{high:,.0f}{suffix}"
        value = low or high
        return f"{symbol}{value:,.0f}{suffix}" if value else ""

    if job.day_rate_min or job.day_rate_max:
        return band(job.day_rate_min, job.day_rate_max, "/day")
    if job.salary_min or job.salary_max:
        return band(job.salary_min, job.salary_max)
    return "Not published"


def region_summary(job: Job) -> str:
    """Condense remote status + eligibility into one scannable word or two."""
    status = job.remote_status or ""
    if job.is_prospect or "Unknown" in status or "unconfirmed" in status.lower():
        return "Verify region"
    if "Worldwide" in status:
        return "Worldwide"
    if "International contractor" in status:
        return "Intl contractor"
    home = profile.active().home_country
    if "regional" in status:
        return "Region (eligible)"
    if home and home in status:
        return home
    return status or "Unknown"


QUICK_COLUMNS: list[tuple[str, Callable[[Job], Any], str]] = [
    ("Company",    lambda j: j.company,                       "text"),
    ("Region",     lambda j: region_summary(j),               "region"),
    ("Shortlist %", lambda j: j.match_score,                  "score"),
    ("Level",      lambda j: j.experience_level,              "level"),
    ("Salary",     lambda j: pay_summary(j),                  "text"),
    ("Email",      lambda j: j.public_email,                  "email"),
    ("Website",    lambda j: j.company_website,               "url"),
    ("Job Link",   lambda j: j.application_url or j.original_job_url, "url"),
]

ELIGIBLE_STYLE = (PatternFill("solid", fgColor="C6EFCE"), Font(bold=True, color="006100"))
INELIGIBLE_STYLE = (PatternFill("solid", fgColor="F8CBAD"), Font(bold=True, color="9C0006"))
UNCONFIRMED_STYLE = (PatternFill("solid", fgColor="FCE4D6"), Font(bold=True, color="9C4500"))


def eligibility_style(value: str):
    """Colour the verdict, not the country it happens to name.

    These strings are written per search — "Eligible — Nigeria named in the
    location" — so they can only be matched on how they open.
    """
    text = (value or "").strip().lower()
    if text.startswith("not eligible"):
        return INELIGIBLE_STYLE
    if text.startswith("eligible") or "confirmed eligible" in text:
        return ELIGIBLE_STYLE
    if text.startswith(("unconfirmed", "unknown")):
        return UNCONFIRMED_STYLE
    return None


def _write_quick_sheet(wb: Workbook, jobs: list[Job]) -> None:
    """A deliberately minimal sheet: company, email, website, link, pay, level."""
    ws = wb.create_sheet("Quick Apply")
    headers = [name for name, _, _ in QUICK_COLUMNS]
    kinds = [kind for _, _, kind in QUICK_COLUMNS]
    _style_header(ws, headers)

    jobs = _sorted(_dedupe(jobs))
    raw_rows: list[list[Any]] = []

    for row_index, job in enumerate(jobs, start=2):
        values: list[Any] = []
        for col_index, (_header, getter, kind) in enumerate(QUICK_COLUMNS, start=1):
            value = getter(job)
            cell = ws.cell(row=row_index, column=col_index)
            values.append(value)

            if kind == "url":
                _hyperlink(cell, value or "")
            elif kind == "email":
                if value:
                    _hyperlink(cell, f"mailto:{value}", value)
                else:
                    cell.value = "—"
            elif kind == "level":
                cell.value = value
                cell.alignment = CENTER
                if value in LEVEL_STYLES:
                    cell.fill, cell.font = LEVEL_STYLES[value]
            elif kind == "region":
                cell.value = value
                cell.alignment = CENTER
                style = eligibility_style(value)
                if style:
                    cell.fill, cell.font = style
            elif kind == "score":
                cell.value = value
                cell.alignment = CENTER
                cell.font = Font(bold=job.match_score >= config.HOT_LEAD_SCORE)
            else:
                cell.value = value
                cell.alignment = TOP

            cell.border = BORDER

            if job.match_score >= config.HOT_LEAD_SCORE and kind in ("text",):
                cell.fill = HOT_FILL
        raw_rows.append(values)

    if not jobs:
        ws.cell(row=2, column=1,
                value="No qualifying opportunities in this run.").font = Font(
                    italic=True, color="808080")

    _autosize(ws, headers, kinds, raw_rows)
    for letter, width in (("A", 30), ("B", 16), ("C", 10), ("D", 15),
                          ("E", 22), ("F", 30), ("G", 32), ("H", 46)):
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"
    _add_table(ws, "QuickApply_tbl", len(jobs) + 1, len(QUICK_COLUMNS))
    ws.sheet_view.showGridLines = False


def _write_companies_sheet(wb: Workbook, companies: list[dict]) -> None:
    ws = wb.create_sheet("Companies & Contacts")
    headers = [h for h, _, _ in COMPANY_COLUMNS]
    kinds = [k for _, _, k in COMPANY_COLUMNS]
    _style_header(ws, headers)

    raw_rows: list[list[Any]] = []
    for row_index, entry in enumerate(companies, start=2):
        values: list[Any] = []
        for col_index, (_header, key, kind) in enumerate(COMPANY_COLUMNS, start=1):
            value = entry.get(key)
            cell = ws.cell(row=row_index, column=col_index)
            values.append(value)
            if kind == "url":
                _hyperlink(cell, value or "")
                cell.alignment = TOP
            elif kind == "email":
                if value:
                    _hyperlink(cell, f"mailto:{value}", value)
                cell.alignment = TOP
            elif kind == "date":
                cell.value = value
                cell.number_format = DATE_FORMAT
                cell.alignment = CENTER
            elif kind == "score":
                cell.value = value
                cell.alignment = CENTER
            elif kind == "wrap":
                cell.value = value
                cell.alignment = WRAP
            else:
                cell.value = value
                cell.alignment = TOP
            cell.border = BORDER
            if entry.get("best_match_score", 0) >= config.HOT_LEAD_SCORE:
                cell.fill = HOT_FILL
        raw_rows.append(values)

    if not companies:
        ws.cell(row=2, column=1, value="No companies to contact in this run.").font = Font(
            italic=True, color="808080")

    _autosize(ws, headers, kinds, raw_rows)
    ws.freeze_panes = "A2"
    _add_table(ws, "Companies_tbl", len(companies) + 1, len(COMPANY_COLUMNS))
    ws.sheet_view.showGridLines = False


def _write_summary_sheet(wb: Workbook, result: RunResult) -> None:
    stats: RunStats = result.stats
    ws = wb.create_sheet("Search Summary")
    ws.sheet_view.showGridLines = False

    search = profile.active()
    title = ws.cell(row=1, column=1,
                    value=f"{search.label} job search — run summary")
    title.font = Font(name="Calibri", size=16, bold=True, color="1F3864")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    subtitle = ws.cell(row=2, column=1, value=_summary_subtitle(search))
    subtitle.font = Font(name="Calibri", size=10, italic=True, color="595959")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    row = 4

    def section(heading: str) -> None:
        nonlocal row
        cell = ws.cell(row=row, column=1, value=heading)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 22
        row += 1

    def metric(label: str, value: Any, *, highlight: bool = False,
               number_format: str | None = None) -> None:
        nonlocal row
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(name="Calibri", size=11, bold=highlight)
        label_cell.alignment = Alignment(vertical="center", indent=1)
        label_cell.border = BORDER

        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = Font(name="Calibri", size=11, bold=True,
                               color="C00000" if highlight else "000000")
        value_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        value_cell.border = BORDER
        if number_format:
            value_cell.number_format = number_format
        if highlight:
            value_cell.fill = HOT_FILL
        row += 1

    period = f"{stats.period_start} to {stats.period_end} ({local_timezone().key})"

    section("Run")
    metric("Run date", stats.run_at.date() if stats.run_at else None, number_format=DATE_FORMAT)
    metric("Run time", stats.run_at.strftime("%H:%M %Z") if stats.run_at else "")
    metric("Search period", period)
    metric("Freshness window", f"Last {stats.freshness_days} calendar days")
    metric("Pay floor", f"${config.SALARY_FLOOR_USD:,.0f} / year USD equivalent")
    metric("Number of sources searched", stats.sources_searched)
    row += 1

    section("Funnel")
    metric("Raw jobs found", stats.raw_found)
    metric("Full adverts fetched for truncated postings", stats.descriptions_filled)
    metric(f"Not {search.label} relevant (rejected)", stats.rejected_irrelevant)
    metric(f"Jobs older than {stats.freshness_days} days rejected", stats.rejected_stale)
    metric("Kept with no published date", stats.undated_kept)
    metric("Non-remote jobs rejected", stats.rejected_not_remote)
    metric("Geographically ineligible jobs rejected", stats.rejected_ineligible)
    metric("Low-rate market rejected", stats.rejected_low_rate_market)
    metric("Large employers rejected", stats.rejected_large_employer)
    metric("Below pay floor rejected", stats.rejected_low_pay)
    metric("Below match threshold (rejected)", stats.rejected_low_score)
    metric("Closed / expired adverts rejected", stats.rejected_expired)
    metric("Duplicates removed", stats.duplicates_removed)
    metric("Qualified jobs", stats.qualified, highlight=True)
    row += 1

    section("Results")
    metric("Hot Leads (score ≥ 85)", stats.hot_leads, highlight=True)
    metric("New since the last run", stats.new_since_last_run)
    metric("Full-time jobs", stats.full_time)
    metric("Part-time jobs", stats.part_time)
    metric("Contract jobs", stats.contract)
    metric("Freelance opportunities", stats.freelance)
    metric("Startup prospects", stats.startups)
    metric("Partnership opportunities", stats.partnerships)
    metric("Prospects (eligibility unconfirmed)", stats.prospects)
    metric("Companies to contact", stats.companies)
    row += 1

    section("Experience level (tailor your CV per level)")
    metric("Beginner / entry level (0-2 yrs)", stats.level_beginner)
    metric("Medium / mid level (3-4 yrs)", stats.level_medium)
    metric("Senior / lead (5+ yrs)", stats.level_senior)
    metric("Level not specified", stats.level_unspecified)
    row += 1

    section("Sources searched")
    header_row = row
    for col, name in enumerate(("Source", "Postings returned", "Status", "Seconds"), start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(name="Calibri", size=11, bold=True, color="1F3864")
        cell.border = BORDER
        cell.alignment = CENTER
    row += 1

    for stat in stats.sources:
        ws.cell(row=row, column=1, value=stat.name).border = BORDER
        ws.cell(row=row, column=2, value=stat.raw_count).border = BORDER
        status_cell = ws.cell(row=row, column=3, value="OK" if stat.ok else f"Failed: {stat.error}")
        status_cell.border = BORDER
        status_cell.font = Font(color="006100" if stat.ok else "9C0006")
        ws.cell(row=row, column=4, value=stat.elapsed).border = BORDER
        row += 1

    row += 1
    section("Filter rules applied")
    for rule in _filter_rules(search, stats.freshness_days):
        cell = ws.cell(row=row, column=1, value="• " + rule)
        cell.font = Font(name="Calibri", size=10, color="404040")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 15
        row += 1

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.freeze_panes = "A4"


def _summary_subtitle(search) -> str:
    """One line describing what this run looked for."""
    reach = "Remote work only." if search.remote_only else "Remote and on-site work."
    if search.home_country:
        return (f"{reach} {search.label} opportunities a resident of "
                f"{search.home_country} is eligible to perform.")
    return f"{reach} {search.label} opportunities."


def _filter_rules(search, freshness_days: int) -> tuple[str, ...]:
    """The rules this run actually applied, in the words of this search."""
    where = search.home_country or "your location"
    if search.remote_only:
        reach = ("REMOTE ONLY — hybrid, on-site, office-based and any required office "
                 "attendance are rejected.")
    else:
        reach = ("REMOTE AND ON-SITE — both are included; on-site roles are judged on "
                 f"whether they are within reach of {where}.")
    return (
        reach,
        f"Eligibility must be positively evidenced: {where} named, worldwide/anywhere, "
        f"a region with {where} explicitly included, or international contractor hiring.",
        "Restrictions to a single other country are rejected.",
        f"'Remote' with no stated region is NOT assumed to be open to {where} — "
        f"it becomes a Prospect.",
        f"Posted within the last {freshness_days} calendar days.",
        "Duplicate vacancies across boards are collapsed to a single row.",
    )


def _write_legend(ws: Worksheet, start_row: int) -> None:
    entries = (
        (f"Match % ≥ {config.HOT_LEAD_SCORE} (Hot Lead)", HOT_FILL),
        ("Job title tinted green — posted today or yesterday", FRESH_FILL),
        ("Job title in bold — NEW, not seen in a previous run", NEW_FILL),
        ("Seen in an earlier run", SEEN_FILL),
        ("Prospect — eligibility unconfirmed", PROSPECT_FILL),
        ("Quick Apply Level: Beginner (0-2 yrs)", LEVEL_STYLES[config.LEVEL_BEGINNER][0]),
        ("Quick Apply Level: Medium (3-4 yrs)", LEVEL_STYLES[config.LEVEL_MEDIUM][0]),
        ("Quick Apply Level: Senior (5+ yrs)", LEVEL_STYLES[config.LEVEL_SENIOR][0]),
    )
    ws.cell(row=start_row, column=1, value="Legend").font = Font(bold=True, size=12, color="1F3864")
    for offset, (label, fill) in enumerate(entries, start=1):
        swatch = ws.cell(row=start_row + offset, column=1)
        swatch.fill = fill
        swatch.border = BORDER
        ws.cell(row=start_row + offset, column=2, value=label).font = Font(size=11)


def build_workbook(result: RunResult, path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    _write_quick_sheet(wb, result.qualified)

    _write_job_sheet(wb, "Hot Leads", result.hot_leads,
                     subtitle=f"No opportunities scored {config.HOT_LEAD_SCORE} or above in this run.")
    _write_job_sheet(wb, "All Qualified Jobs", result.qualified)
    _write_job_sheet(wb, "Full Time", result.full_time)
    _write_job_sheet(wb, "Part Time", result.part_time)
    _write_job_sheet(wb, "Contract", result.contract)
    _write_job_sheet(wb, "Freelance", result.freelance)
    _write_job_sheet(wb, "Startups", result.startups)
    _write_job_sheet(wb, "Partnerships", result.partnerships)
    _write_job_sheet(wb, "Prospects", result.prospects,
                     subtitle="Leads whose eligibility could not be verified — ask before applying.",
                     columns=PROSPECT_COLUMNS)
    _write_job_sheet(wb, "Long Shots", result.long_shots,
                     subtitle="Qualified on every rule, but rated a low chance of a reply "
                              "against your CV. Worth a look when the sheets above are thin.")
    _write_companies_sheet(wb, result.companies)
    _write_summary_sheet(wb, result)

    summary = wb["Search Summary"]
    _write_legend(summary, summary.max_row + 2)

    search = profile.active()
    wb.properties.title = f"{search.label} opportunities"
    wb.properties.creator = "Job Finder"
    wb.properties.description = (
        f"{_summary_subtitle(search)} Posted in the last "
        f"{result.stats.freshness_days} days."
    )

    defuse_formulas(wb)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
