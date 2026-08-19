"""Tests for de-duplication, compensation parsing and the Excel report."""

from __future__ import annotations

import csv
import io
import json
import tempfile
import unittest
from contextlib import redirect_stdout
from datetime import timedelta
from pathlib import Path

import openpyxl

from job_agent import config, pipeline, profile
from job_agent.dedupe import deduplicate, fingerprint
from job_agent.models import Job, RawJob
from job_agent.report_data import write_csv, write_json
from job_agent.report_excel import COLUMNS, build_workbook, eligibility_style
from job_agent.utils import now_local, parse_compensation, parse_datetime

from .fixtures import flutter_uk_profile

EXPECTED_SHEETS = [
    "Quick Apply", "Hot Leads", "All Qualified Jobs", "Full Time", "Part Time", "Contract",
    "Freelance", "Startups", "Partnerships", "Prospects", "Long Shots",
    "Companies & Contacts", "Search Summary",
]

EXPECTED_COLUMNS = [
    "Shortlist %", "CV Fit %", "Job Title", "Company", "Location", "Arrangement",
    "Employment", "Salary", "Posted", "Potential gaps", "Email", "Contact",
    "Website", "Job Link", "Why this rank",
]


class CompensationTests(unittest.TestCase):

    def test_annual_salary_range(self):
        parsed = parse_compensation("£75,000 - £95,000 per annum")
        self.assertEqual(parsed["salary_min"], 75_000)
        self.assertEqual(parsed["salary_max"], 95_000)
        self.assertEqual(parsed["currency"], "GBP")

    def test_k_suffix(self):
        parsed = parse_compensation("$90k - $120k")
        self.assertEqual(parsed["salary_min"], 90_000)
        self.assertEqual(parsed["salary_max"], 120_000)

    def test_day_rate(self):
        parsed = parse_compensation("£475 - £550 per day")
        self.assertEqual(parsed["day_rate_min"], 475)
        self.assertEqual(parsed["day_rate_max"], 550)
        self.assertIsNone(parsed["salary_min"])

    def test_hourly_converted_to_day_rate(self):
        parsed = parse_compensation("$70 - $95 per hour")
        self.assertEqual(parsed["day_rate_min"], 560)
        self.assertEqual(parsed["day_rate_max"], 760)

    def test_years_are_not_salaries(self):
        parsed = parse_compensation("Founded in 2019, we now have a big team")
        self.assertIsNone(parsed["salary_min"])


class DateTests(unittest.TestCase):

    def test_epoch_seconds(self):
        self.assertEqual(parse_datetime(1786354319).year, 2026)

    def test_iso_and_rfc822(self):
        self.assertIsNotNone(parse_datetime("2026-08-05T09:30:00Z"))
        self.assertIsNotNone(parse_datetime("Sat, 08 Aug 2026 02:47:00 +0000"))

    def test_javascript_date_tostring(self):
        parsed = parse_datetime("Tue Jul 21 2026 22:30:40 GMT+0000 (Coordinated Universal Time)")
        self.assertIsNotNone(parsed)
        self.assertEqual(parsed.date().isoformat(), "2026-07-21")

    def test_relative(self):
        parsed = parse_datetime("3 days ago")
        self.assertEqual((now_local().date() - parsed.date()).days, 3)

    def test_junk_returns_none(self):
        self.assertIsNone(parse_datetime("whenever"))


def make_job(company="Acme", title="Senior Flutter Developer", score=90,
             days_ago=0, **kwargs) -> Job:
    posted = now_local() - timedelta(days=days_ago)
    job = Job(
        fingerprint=fingerprint(company, title),
        company=company, title=title, match_score=score,
        posted_at=posted, posted_date=posted.date(),
        job_age_days=days_ago, discovered_date=posted.date(),
        verified_date=posted.date(),
        employment_type="Full Time", remote_status="Remote — UK",
        eligibility="Eligible", source="test",
    )
    for key, value in kwargs.items():
        setattr(job, key, value)
    return job


class DedupeTests(unittest.TestCase):

    def test_same_role_across_two_boards_is_merged(self):
        a = make_job(source="remoteok", description="short")
        b = make_job(title="Senior Flutter Developer (Remote)", source="remotive",
                     description="a much longer description", public_email="jobs@acme.com")
        merged, removed = deduplicate([a, b])
        self.assertEqual(removed, 1)
        self.assertEqual(len(merged), 1)
        self.assertEqual(merged[0].public_email, "jobs@acme.com",
                         "contact details found by either copy must survive")

    def test_different_companies_are_kept(self):
        merged, removed = deduplicate([make_job(company="Acme"), make_job(company="Globex")])
        self.assertEqual(removed, 0)
        self.assertEqual(len(merged), 2)

    def test_distinct_jobs_at_one_ats_board_are_not_merged(self):
        """Regression: numeric ids on ATS URLs identify the vacancy. Stripping them
        collapsed every job at a company into a single row.
        """
        a = make_job(title="Web Frontend Engineer",
                     original_job_url="https://job-boards.greenhouse.io/canonical/jobs/5150422")
        b = make_job(title="Ubuntu Engineering Lead",
                     original_job_url="https://job-boards.greenhouse.io/canonical/jobs/5150999")
        merged, removed = deduplicate([a, b])
        self.assertEqual(removed, 0, "two different vacancies must both survive")
        self.assertEqual(len(merged), 2)

    def test_same_vacancy_same_url_is_merged(self):
        url = "https://job-boards.greenhouse.io/canonical/jobs/5150422"
        a = make_job(title="Web Frontend Engineer", original_job_url=url)
        b = make_job(title="Web Frontend Engineer (Remote)", original_job_url=url)
        merged, removed = deduplicate([a, b])
        self.assertEqual(removed, 1)

    def test_agency_reference_numbers_do_not_split_one_vacancy(self):
        jobs = [
            make_job(company="BairesDev", title="Senior Flutter Engineer | REF#300611"),
            make_job(company="BairesDev", title="Senior Flutter Engineer | REF#412998"),
        ]
        merged, removed = deduplicate(jobs)
        self.assertEqual(len(merged), 1)
        self.assertEqual(removed, 1)

    def test_company_tagline_does_not_split_one_vacancy(self):
        jobs = [
            make_job(company="Big Potato", title="Flutter Developer"),
            make_job(company="Big Potato | Board Game Company", title="Flutter Developer"),
        ]
        self.assertEqual(len(deduplicate(jobs)[0]), 1)

    def test_longer_company_name_merges_into_the_shorter_one(self):
        jobs = [
            make_job(company="Big Potato Games", title="Flutter Developer"),
            make_job(company="Big Potato", title="Flutter Developer"),
        ]
        self.assertEqual(len(deduplicate(jobs)[0]), 1)

    def test_similar_names_with_different_roles_stay_apart(self):
        jobs = [
            make_job(company="Big Potato Games", title="Flutter Developer"),
            make_job(company="Big Potato", title="Android Developer"),
        ]
        self.assertEqual(len(deduplicate(jobs)[0]), 2)

    def test_short_company_stems_never_merge_on_a_coincidence(self):
        jobs = [
            make_job(company="Arc", title="Flutter Developer"),
            make_job(company="Arcadia Software", title="Flutter Developer"),
        ]
        self.assertEqual(len(deduplicate(jobs)[0]), 2)

    def test_earliest_posted_date_wins(self):
        older = make_job(days_ago=4, source="a")
        newer = make_job(days_ago=1, source="b", description="longer text " * 20)
        merged, _ = deduplicate([newer, older])
        self.assertEqual(merged[0].posted_date, older.posted_date)


class ExcelReportTests(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        jobs = [
            make_job(company="Hot Co", score=92, days_ago=0, is_new=True,
                     salary_min=80000, salary_max=100000, salary_currency="GBP",
                     application_url="https://hotco.com/apply",
                     public_email="careers@hotco.com"),
            make_job(company="Mid Co", score=70, days_ago=3, is_new=False,
                     employment_type="Contract", contract_type="Outside IR35",
                     day_rate_min=500, day_rate_max=600, salary_currency="GBP"),
            make_job(company="Free Co", score=65, days_ago=2,
                     employment_type="Freelance", is_startup=True),
        ]
        result = pipeline.RunResult(stats=pipeline.RunStats(run_at=now_local()))
        result.qualified = jobs
        result.prospects = []
        pipeline._bucket(result)
        cls.result = result
        cls.tmp = tempfile.TemporaryDirectory()
        cls.path = build_workbook(result, Path(cls.tmp.name) / "test.xlsx")
        cls.wb = openpyxl.load_workbook(cls.path)

    @classmethod
    def tearDownClass(cls):
        cls.tmp.cleanup()

    def test_all_sheets_present_in_order(self):
        self.assertEqual(self.wb.sheetnames, EXPECTED_SHEETS)

    def test_column_schema_exact(self):
        headers = [c.value for c in self.wb["All Qualified Jobs"][1]]
        self.assertEqual(headers, EXPECTED_COLUMNS)
        self.assertEqual(len(COLUMNS), len(EXPECTED_COLUMNS))

    def test_header_row_is_frozen_on_every_job_sheet(self):
        for name in EXPECTED_SHEETS[1:10]:
            self.assertEqual(self.wb[name].freeze_panes, "A2", name)

    def test_tables_and_filters_exist(self):
        ws = self.wb["All Qualified Jobs"]
        self.assertTrue(ws.tables, "expected an Excel table providing filters")

    def test_hot_leads_only_contains_high_scores(self):
        ws = self.wb["Hot Leads"]
        scores = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
        self.assertTrue(scores, "expected at least one hot lead")
        self.assertTrue(all(s >= config.HOT_LEAD_SCORE for s in scores), scores)

    def test_rows_sorted_by_score_desc(self):
        ws = self.wb["All Qualified Jobs"]
        scores = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
        self.assertEqual(scores, sorted(scores, reverse=True))

    def test_salary_is_one_readable_summary_not_five_columns(self):
        ws = self.wb["All Qualified Jobs"]
        value = ws.cell(2, EXPECTED_COLUMNS.index("Salary") + 1).value
        self.assertTrue(value, "every row states its pay or says it is not published")
        self.assertTrue(value.startswith("£") or value == "Not published", value)

    def test_urls_are_clickable(self):
        ws = self.wb["All Qualified Jobs"]
        cell = ws.cell(2, EXPECTED_COLUMNS.index("Job Link") + 1)
        self.assertIsNotNone(cell.hyperlink)
        self.assertEqual(cell.hyperlink.target, "https://hotco.com/apply")

    def test_csv_keeps_every_field_the_workbook_drops(self):
        """The narrow workbook is only safe because the CSV stays complete."""
        import csv as csv_mod
        from job_agent.report_data import write_csv

        path = Path(self.tmp.name) / "full.csv"
        write_csv(self.result, path)
        with path.open(encoding="utf-8-sig") as handle:
            headers = next(csv_mod.reader(handle))
        for field in ("remote_status", "eligibility", "concerns", "match_reasons",
                      "seniority", "industry", "posted_date", "networking_score",
                      "application_url", "original_job_url"):
            self.assertIn(field, headers)

    def test_prospects_sheet_says_why_it_is_only_a_prospect(self):
        ws = self.wb["Prospects"]
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers, EXPECTED_COLUMNS + ["Why unconfirmed"])

    def test_hot_row_is_highlighted(self):
        ws = self.wb["All Qualified Jobs"]
        # Not the title cell: a fresh posting tints that green instead.
        company_col = EXPECTED_COLUMNS.index("Company") + 1
        self.assertEqual(ws.cell(2, company_col).fill.fgColor.rgb, "00FFF2CC")

    def test_summary_reports_the_funnel(self):
        ws = self.wb["Search Summary"]
        labels = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, ws.max_row + 1)}
        for required in ("Raw jobs found",
                         f"Jobs older than {config.FRESHNESS_DAYS} days rejected",
                         "Non-remote jobs rejected", "Geographically ineligible jobs rejected",
                         "Large employers rejected",
                         "Duplicates removed", "Qualified jobs",
                         f"Hot Leads (score ≥ {config.HOT_LEAD_SCORE})",
                         "Full-time jobs", "Part-time jobs", "Contract jobs",
                         "Freelance opportunities", "Startup prospects",
                         "Partnership opportunities", "Search period",
                         "Number of sources searched"):
            self.assertIn(required, labels)
        self.assertEqual(labels["Qualified jobs"], 3)

    def test_file_is_a_real_xlsx_not_a_renamed_csv(self):
        with open(self.path, "rb") as handle:
            self.assertEqual(handle.read(2), b"PK", "xlsx must be a ZIP container")


if __name__ == "__main__":
    unittest.main(verbosity=2)


class QuickModeTests(unittest.TestCase):
    """Quick mode must actually shrink the budgets it claims to."""

    FIELDS = ("LINKEDIN_MAX_DETAILS", "DESCRIPTION_MAX_FETCHES",
              "LINKEDIN_SEARCH_WORKERS", "LLM_MAX_ELIGIBILITY_CALLS",
              "LLM_MAX_FIT_CALLS", "LLM_CONCURRENCY")

    def setUp(self):
        self._saved = {f: getattr(config, f) for f in self.FIELDS}

    def tearDown(self):
        for field, value in self._saved.items():
            setattr(config, field, value)

    def _linkedin_budget(self):
        from job_agent.sources.linkedin import LinkedIn
        source = LinkedIn()
        return (source.MAX_DETAILS if source.MAX_DETAILS is not None
                else config.LINKEDIN_MAX_DETAILS)

    def test_the_budgets_shrink(self):
        before = self._linkedin_budget()
        config.use_quick_mode()
        self.assertLess(self._linkedin_budget(), before)
        self.assertEqual(self._linkedin_budget(), config.QUICK_LINKEDIN_MAX_DETAILS)
        self.assertEqual(config.DESCRIPTION_MAX_FETCHES,
                         config.QUICK_DESCRIPTION_MAX_FETCHES)

    def test_the_judgement_layer_is_bounded_too(self):
        before = config.LLM_MAX_ELIGIBILITY_CALLS
        config.use_quick_mode()
        self.assertLess(config.LLM_MAX_ELIGIBILITY_CALLS, before)
        self.assertEqual(config.LLM_MAX_FIT_CALLS, config.QUICK_LLM_MAX_FIT_CALLS)
        self.assertGreater(config.LLM_CONCURRENCY, 4)

    def test_the_linkedin_budget_is_read_at_run_time_not_import_time(self):
        config.LINKEDIN_MAX_DETAILS = 7
        self.assertEqual(self._linkedin_budget(), 7)

    def test_quick_mode_is_reversible(self):
        before = {name: getattr(config, name) for name, _ in config._QUICK_FIELDS}
        saved = config.use_quick_mode()
        self.assertNotEqual(config.LINKEDIN_MAX_DETAILS, before["LINKEDIN_MAX_DETAILS"])
        config.restore_budgets(saved)
        for name, value in before.items():
            self.assertEqual(getattr(config, name), value, name)

    def test_an_explicit_per_instance_override_still_wins(self):
        from job_agent.sources.linkedin import LinkedIn
        source = LinkedIn()
        source.MAX_DETAILS = 3
        self.assertEqual(source.MAX_DETAILS, 3)


class EffectivePayFloorTests(unittest.TestCase):
    """The run must report the floor it used, not the one it was asked for."""

    def tearDown(self):
        from job_agent import profile
        profile.reset()

    def test_a_profile_floor_overrides_the_default_and_is_reported(self):
        from dataclasses import replace
        custom = replace(flutter_uk_profile(), key="compiled:test", salary_floor_usd=85_000)
        result = pipeline.run(offline=True, verify_live=False, use_llm=False,
                              search_profile=custom)
        self.assertEqual(result.stats.salary_floor_usd, 85_000)

    def test_an_explicit_floor_is_not_overridden(self):
        from dataclasses import replace
        custom = replace(flutter_uk_profile(), key="compiled:test", salary_floor_usd=85_000)
        result = pipeline.run(offline=True, verify_live=False, use_llm=False,
                              min_salary_usd=120_000, search_profile=custom)
        self.assertEqual(result.stats.salary_floor_usd, 120_000)


class StatedFloorRoutingTests(unittest.TestCase):
    """A stated minimum turns "pay not published" into a question, not a verdict."""

    def tearDown(self):
        from job_agent import profile
        profile.reset()

    def _run(self, *, stated: bool):
        from dataclasses import replace
        custom = replace(flutter_uk_profile(), key="compiled:test",
                         salary_floor_usd=50_000, pay_floor_stated=stated)
        return pipeline.run(offline=True, verify_live=False, use_llm=False,
                            search_profile=custom)

    def test_unpublished_pay_becomes_a_prospect_when_a_floor_was_stated(self):
        result = self._run(stated=True)
        self.assertGreater(result.stats.pay_unstated, 0)
        moved = [j for j in result.prospects if j.rejection_category == "pay_unstated"]
        self.assertTrue(moved)
        self.assertTrue(moved[0].is_prospect)

    def test_the_row_says_what_to_check(self):
        moved = [j for j in self._run(stated=True).prospects
                 if j.rejection_category == "pay_unstated"]
        self.assertEqual(moved[0].application_status, "Ask about pay before applying")

    def test_an_estimated_floor_leaves_them_qualified(self):
        result = self._run(stated=False)
        self.assertEqual(result.stats.pay_unstated, 0)

    def test_a_stated_floor_costs_qualified_leads_not_hides_them(self):
        stated, estimated = self._run(stated=True), self._run(stated=False)
        self.assertLess(stated.stats.qualified, estimated.stats.qualified)
        self.assertEqual(stated.stats.qualified + stated.stats.prospects,
                         estimated.stats.qualified + estimated.stats.prospects)


class ProspectRoutingByCategoryTests(unittest.TestCase):
    """keep_prospect must read the category, not sniff the reason text."""

    def tearDown(self):
        from job_agent import profile
        profile.reset()

    @staticmethod
    def _job(category: str, reason: str, score: float):
        from job_agent.models import Job
        job = Job(title="Site Engineer", company="Acme")
        job.rejection_category = category
        job.rejection_reason = reason
        job.match_score = score
        return job

    def test_unpublished_pay_is_judged_on_the_ordinary_qualifying_bar(self):
        from job_agent import config, pipeline
        job = self._job("pay_unstated", "Salary not stated in the advert",
                        config.MIN_QUALIFY_SCORE)
        self.assertTrue(pipeline.keep_prospect(job))

    def test_unpublished_pay_below_the_qualifying_bar_is_dropped(self):
        from job_agent import config, pipeline
        job = self._job("pay_unstated", "Salary not stated in the advert",
                        config.MIN_QUALIFY_SCORE - 1)
        self.assertFalse(pipeline.keep_prospect(job))

    def test_a_pay_reason_is_not_scored_against_the_eligibility_bar(self):
        from job_agent import config, pipeline
        self.assertLess(config.PROSPECT_UNVERIFIED_MIN_SCORE, config.MIN_QUALIFY_SCORE)
        job = self._job("pay_unstated", "Pay is not stated",
                        config.PROSPECT_UNVERIFIED_MIN_SCORE)
        self.assertFalse(pipeline.keep_prospect(job))

    def test_undetermined_eligibility_keeps_the_lower_bar(self):
        from job_agent import config, pipeline
        job = self._job("ineligible", "Location is not confirmed",
                        config.PROSPECT_UNVERIFIED_MIN_SCORE)
        self.assertTrue(pipeline.keep_prospect(job))

    def test_a_core_term_in_the_title_carries_an_undetermined_advert(self):
        from dataclasses import replace
        from job_agent import pipeline, profile
        profile.set_active(replace(profile.active(), key="compiled:test",
                                   core_terms=("plumbing",)))
        job = self._job("ineligible", "Location is not confirmed", 1.0)
        job.title = "Plumbing Supervisor"
        self.assertTrue(pipeline.keep_prospect(job))


class LongShotVisibilityTests(unittest.TestCase):
    """A lead the model rates a long shot is set aside, not thrown away.

    It has already passed relevance, freshness, eligibility, pay and size —
    on a thin search it may be the best thing available, so it stays in the
    machine-readable exports. The workbook keeps its twelve sheets.
    """

    def setUp(self):
        self.result = pipeline.RunResult(stats=pipeline.RunStats(run_at=now_local()))
        self.result.qualified = [make_job(company="Strong Co", score=80)]
        self.result.long_shots = [make_job(company="Long Shot Co", score=30)]
        pipeline._bucket(self.result)
        self.tmp = tempfile.TemporaryDirectory()
        self.dir = Path(self.tmp.name)

    def tearDown(self):
        self.tmp.cleanup()

    def test_the_json_export_carries_them_and_counts_them(self):
        path = write_json(self.result, self.dir / "r.json")
        payload = json.loads(path.read_text(encoding="utf-8"))
        self.assertEqual(payload["counts"]["long_shots"], 1)
        self.assertEqual([j["company"] for j in payload["long_shots"]], ["Long Shot Co"])

    def test_the_csv_labels_them_rather_than_dropping_them(self):
        path = write_csv(self.result, self.dir / "r.csv")
        with path.open(encoding="utf-8-sig") as handle:
            rows = list(csv.DictReader(handle))
        sheets = {row["Sheet"] for row in rows}
        self.assertIn("Long Shot", sheets)
        self.assertEqual(len(rows), 2)

    def test_they_get_their_own_sheet_without_disturbing_the_others(self):
        # The workbook is compared against the previous run, so the sheet list
        # is a contract: "Long Shots" is added, nothing is renamed or reordered.
        path = build_workbook(self.result, self.dir / "r.xlsx")
        wb = openpyxl.load_workbook(path)
        self.assertEqual(wb.sheetnames, EXPECTED_SHEETS)
        self.assertEqual([c.value for c in wb["Long Shots"][1]], EXPECTED_COLUMNS)
        self.assertEqual(wb["Long Shots"]["D2"].value, "Long Shot Co")


class EligibilityStylingTests(unittest.TestCase):
    """The eligibility column is coloured by its verdict.

    It was keyed on fixed strings ("UK", "Europe (UK ok)") that no longer
    matched anything the filters wrote, so the column had been unstyled for
    every search regardless of country.
    """

    def test_a_verdict_is_styled_whatever_country_it_names(self):
        for value in ("Eligible — Nigeria named in the location",
                      "Eligible — worldwide / work from anywhere",
                      "Not eligible — US residency/authorisation required",
                      "Unconfirmed — ask before applying",
                      "Unknown — set a region or supply a CV"):
            with self.subTest(value=value):
                self.assertIsNotNone(eligibility_style(value))

    def test_eligible_and_ineligible_do_not_share_a_colour(self):
        eligible = eligibility_style("Eligible — Nigeria named in the location")
        refused = eligibility_style("Not eligible — US residency required")
        self.assertNotEqual(eligible[0].fgColor.rgb, refused[0].fgColor.rgb)

    def test_a_value_that_states_nothing_is_left_plain(self):
        self.assertIsNone(eligibility_style("N/A"))
        self.assertIsNone(eligibility_style(""))


class CheckMirrorsTheRunTests(unittest.TestCase):
    """`jobfinder check` must explain the run, not a stricter rule of its own.

    `assess_remote` reads only the advert; whether its arrangement disqualifies
    it is a question about the search. The screen settles that through
    `judge_arrangement`, and `check` used to skip the step — so an on-site job
    that a local search would happily keep was reported as "REJECT — no remote
    working statement found".
    """

    def tearDown(self):
        profile.reset()

    def _advert(self):
        return RawJob(source="manual", source_id="m1", title="Electrician",
                      company="Acme", url="https://example.com/job",
                      description="Electrical maintenance and wiring. Full time, on site.",
                      location_raw="Lagos, Nigeria", posted_at=now_local())

    def _searching_for(self, arrangement: str):
        """The gate reads the active profile, so that is what a test must set.

        The advert is in Lagos, so the search is run from Lagos: this class is
        about the arrangement question, and a cross-border on-site role would
        now be rejected on geography before the arrangement was ever reached.
        """
        from dataclasses import replace
        profile.set_active(replace(profile.active(), key="test",
                                   work_arrangement=arrangement,
                                   home_country="Nigeria",
                                   home_terms=("nigeria", "ng"),
                                   home_city_terms=("lagos", "abuja")))

    def test_an_on_site_advert_survives_a_search_that_never_asked_for_remote(self):
        from job_agent.pipeline import judge_arrangement
        self._searching_for("any")
        verdict = judge_arrangement(self._advert(), "any")
        self.assertTrue(verdict.passed, verdict.reason)

    def test_a_search_that_wants_remote_still_rejects_an_on_site_advert(self):
        from job_agent.pipeline import judge_arrangement
        self._searching_for("remote")
        verdict = judge_arrangement(self._advert(), "remote")
        self.assertFalse(verdict.passed)
        self.assertEqual(verdict.category, "wrong_arrangement")

    def test_check_reports_what_the_screen_would_decide(self):
        from job_agent import __main__ as cli
        from job_agent.pipeline import judge_arrangement
        advert = self._advert()
        wanted = profile.active().work_arrangement
        expected = judge_arrangement(advert, wanted)
        args = cli.build_parser().parse_args(
            ["check", "--title", advert.title, "--location", advert.location_raw,
             "--description", advert.description])
        with redirect_stdout(io.StringIO()) as buffer:
            code = args.func(args)
        self.assertEqual(code, 0 if expected.passed else 1)
        self.assertIn("PASS" if expected.passed else "REJECT", buffer.getvalue())


class SpreadsheetsShowTextRatherThanRunItTests(unittest.TestCase):
    """Adverts are third-party text, and the report is opened in Excel.

    A company field of `=cmd|'/c calc'!A1` was written as a live formula,
    which is the standard spreadsheet-injection payload. The characters must
    survive; the behaviour must not.
    """

    PAYLOADS = ("=cmd|'/c calc'!A1", '+HYPERLINK("http://evil","x")',
                "@SUM(1+1)", "-2+3")

    def tearDown(self):
        profile.reset()

    def _report(self, tmp, **kw):
        profile.set_active(flutter_uk_profile())
        from tests.test_pipeline import make_job
        result = pipeline.RunResult(stats=pipeline.RunStats(run_at=now_local()))
        result.qualified = [make_job(score=70, days_ago=1, **kw)]
        result.prospects = []
        pipeline._bucket(result)
        return result

    def test_no_cell_in_the_workbook_is_a_formula(self):
        for payload in self.PAYLOADS:
            with self.subTest(payload=payload), tempfile.TemporaryDirectory() as tmp:
                result = self._report(tmp, company=payload, title=payload)
                path = build_workbook(result, Path(tmp) / "t.xlsx")
                wb = openpyxl.load_workbook(path)
                formulas = [c.coordinate for name in wb.sheetnames
                            for row in wb[name].iter_rows() for c in row
                            if c.data_type == "f"]
                self.assertEqual(formulas, [], f"{payload!r} became a formula")

    def test_the_text_itself_is_kept_exactly(self):
        payload = self.PAYLOADS[0]
        with tempfile.TemporaryDirectory() as tmp:
            result = self._report(tmp, company=payload)
            wb = openpyxl.load_workbook(build_workbook(result, Path(tmp) / "t.xlsx"))
            found = [c.value for name in wb.sheetnames
                     for row in wb[name].iter_rows() for c in row
                     if isinstance(c.value, str) and c.value == payload]
            self.assertTrue(found, "the employer name must still read the same")

    def test_the_csv_does_not_hand_excel_a_formula(self):
        from job_agent.report_data import defuse
        for payload in self.PAYLOADS:
            with self.subTest(payload=payload):
                self.assertTrue(defuse(payload).startswith("'"))

    def test_a_negative_number_is_left_alone(self):
        from job_agent.report_data import defuse
        for number in ("-500", "-1234.56", "-0.5"):
            with self.subTest(number=number):
                self.assertEqual(defuse(number), number)

    def test_ordinary_text_is_untouched(self):
        from job_agent.report_data import defuse
        for text in ("Acme Ltd", "Electrician", "", "A-B Corp"):
            with self.subTest(text=text):
                self.assertEqual(defuse(text), text)

    def test_a_bulleted_description_is_not_quoted(self):
        """Adverts routinely open with "- Qualifications"; that is prose, not a formula."""
        from job_agent.report_data import defuse
        for prose in ("- Qualification Relevent Education", "+ Bonus scheme",
                      "-  double spaced bullet"):
            with self.subTest(prose=prose):
                self.assertEqual(defuse(prose), prose)

    def test_an_equals_or_at_sign_is_quoted_even_before_a_space(self):
        """Neither ever begins ordinary prose, so neither gets the benefit of the doubt."""
        from job_agent.report_data import defuse
        for payload in ("= 1+1", "@ SUM(1)"):
            with self.subTest(payload=payload):
                self.assertTrue(defuse(payload).startswith("'"))

    def test_a_phone_number_is_still_quoted(self):
        """+971 ... is not prose and Excel would mangle it into an error."""
        from job_agent.report_data import defuse
        self.assertEqual(defuse("+971 52 977 3887"), "'+971 52 977 3887")


class ReportsOnlyLinkToAddressesTheyShouldTests(unittest.TestCase):
    """An advert brings its own apply link, and the report makes it clickable.

    A posting offering `javascript:...` had the HTML report render it as an
    "Apply" button and the workbook attach it as a hyperlink, so the reader was
    invited to click something the advert chose. Only web and mail addresses
    are linked now; anything else is shown as plain text.
    """

    UNSAFE = ("javascript:alert(1)", "JavaScript:alert(1)",
              "data:text/html,<script>alert(1)</script>", "file:///etc/passwd")

    def tearDown(self):
        profile.reset()

    def _report(self, url):
        from tests.test_pipeline import make_job
        profile.set_active(flutter_uk_profile())
        result = pipeline.RunResult(stats=pipeline.RunStats(run_at=now_local()))
        result.qualified = [make_job(score=70, days_ago=1, application_url=url)]
        result.prospects = []
        pipeline._bucket(result)
        return result

    def test_safe_url_passes_web_and_mail_addresses(self):
        from job_agent.utils import safe_url
        for good in ("https://example.com/j", "http://example.com", "mailto:a@b.c"):
            with self.subTest(url=good):
                self.assertEqual(safe_url(good), good)

    def test_safe_url_refuses_everything_else(self):
        from job_agent.utils import safe_url
        for bad in self.UNSAFE:
            with self.subTest(url=bad):
                self.assertEqual(safe_url(bad), "")

    def test_the_html_report_does_not_link_them(self):
        from job_agent.report_html import build_html
        for bad in self.UNSAFE:
            with self.subTest(url=bad), tempfile.TemporaryDirectory() as tmp:
                path = Path(tmp) / "r.html"
                build_html(self._report(bad), path)
                self.assertNotIn(f'href="{bad}', path.read_text(encoding="utf-8"))

    def test_the_workbook_does_not_link_them(self):
        for bad in self.UNSAFE:
            with self.subTest(url=bad), tempfile.TemporaryDirectory() as tmp:
                path = build_workbook(self._report(bad), Path(tmp) / "r.xlsx")
                wb = openpyxl.load_workbook(path)
                targets = [c.hyperlink.target for name in wb.sheetnames
                           for row in wb[name].iter_rows() for c in row if c.hyperlink]
                self.assertEqual([t for t in targets if not t.lower().startswith(
                    ("http://", "https://", "mailto:"))], [])

    def test_a_real_apply_link_still_works(self):
        from job_agent.report_html import build_html
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "r.html"
            build_html(self._report("https://real.example/apply"), path)
            self.assertIn('href="https://real.example/apply"', path.read_text(encoding="utf-8"))


class VacancyHistorySurvivesTests(unittest.TestCase):
    """Losing the history must cost the NEW markers, never the report.

    `seen_jobs.sqlite3` only remembers which vacancies have been seen before.
    A corrupt file — an interrupted write, a sync conflict — used to end the
    run in a raw `sqlite3.DatabaseError` traceback with no report written at
    all, throwing away every source fetched.
    """

    def setUp(self):
        from job_agent import config, storage
        self.storage = storage
        self.tmp = tempfile.TemporaryDirectory()
        self.original = config.SEEN_DB
        config.SEEN_DB = Path(self.tmp.name) / "seen_jobs.sqlite3"

    def tearDown(self):
        from job_agent import config
        config.SEEN_DB = self.original
        self.tmp.cleanup()

    def _job(self):
        from tests.test_pipeline import make_job
        return make_job(score=70, days_ago=1)

    def test_a_healthy_history_records_normally(self):
        from job_agent.config import SEEN_DB
        count = self.storage.mark_new_and_record([self._job()], now_local().date())
        self.assertEqual(count, 1)
        self.assertTrue(SEEN_DB.exists())

    def test_a_corrupt_history_is_set_aside_and_rebuilt(self):
        from job_agent import config
        config.SEEN_DB.write_bytes(b"this is not a database")
        count = self.storage.mark_new_and_record([self._job()], now_local().date())
        self.assertEqual(count, 1)
        self.assertTrue(config.SEEN_DB.with_suffix(
            config.SEEN_DB.suffix + ".unreadable").exists(),
            "the unreadable file is kept rather than deleted")

    def test_a_rejected_history_file_is_not_left_open(self):
        """A file that fails the schema must not still be held open.

        `connect` succeeds on any file and the schema is what rejects one that
        is not a database, so the connection is open when that raises. Windows
        will not rename a file another handle holds, which turned "set the
        corrupt history aside and rebuild it" into "this run will not remember
        what it saw" on every Windows runner.
        """
        import sqlite3
        from unittest.mock import MagicMock, patch
        conn = MagicMock()
        conn.executescript.side_effect = sqlite3.DatabaseError("file is not a database")
        with patch.object(self.storage.sqlite3, "connect", return_value=conn), \
                self.assertRaises(sqlite3.DatabaseError):
            self.storage._open(config.SEEN_DB)
        conn.close.assert_called_once()

    def test_an_unwritable_history_still_reports_every_lead(self):
        from unittest.mock import patch
        import sqlite3
        job = self._job()
        with patch.object(self.storage, "_record",
                          side_effect=sqlite3.OperationalError("readonly database")):
            count = self.storage.mark_new_and_record([job], now_local().date())
        self.assertEqual(count, 1)
        self.assertTrue(job.is_new, "with no history, every lead reads as new")
        self.assertEqual(job.job_status, "NEW")

    def test_statistics_degrade_rather_than_raise(self):
        from unittest.mock import patch
        import sqlite3
        with patch.object(self.storage, "_stats",
                          side_effect=sqlite3.DatabaseError("file is not a database")):
            self.assertEqual(self.storage.stats(), {"total_tracked": 0, "applied": 0})

    def test_saving_a_status_degrades_rather_than_raises(self):
        from unittest.mock import patch
        import sqlite3
        with patch.object(self.storage, "_set_status",
                          side_effect=sqlite3.OperationalError("readonly database")):
            self.assertFalse(self.storage.set_application_status("fp", "Applied"))


class SummaryWindowTests(unittest.TestCase):
    """The Search Summary has to describe the run that happened.

    Every one of these lines used to be written from the 30-day default, so
    a 7-day run produced a workbook claiming it had read a month.
    """

    @classmethod
    def setUpClass(cls):
        result = pipeline.RunResult(
            stats=pipeline.RunStats(run_at=now_local(), freshness_days=7,
                                    rejected_stale=4))
        pipeline._bucket(result)
        cls.tmp = tempfile.TemporaryDirectory()
        cls.wb = openpyxl.load_workbook(
            build_workbook(result, Path(cls.tmp.name) / "window.xlsx"))

    @classmethod
    def tearDownClass(cls):
        cls.tmp.cleanup()

    def _summary_text(self) -> str:
        return "\n".join(str(cell.value) for row in self.wb["Search Summary"].iter_rows()
                          for cell in row if cell.value is not None)

    def test_the_window_is_the_one_the_run_used(self):
        self.assertIn("Last 7 calendar days", self._summary_text())

    def test_the_stale_counter_is_labelled_with_that_window(self):
        self.assertIn("Jobs older than 7 days rejected", self._summary_text())

    def test_the_filter_rules_quote_that_window(self):
        self.assertIn("Posted within the last 7 calendar days.", self._summary_text())

    def test_the_workbook_description_carries_that_window(self):
        self.assertIn("Posted in the last 7 days.", self.wb.properties.description)

    def test_no_line_still_claims_the_default(self):
        text = self._summary_text()
        self.assertNotIn(f"Last {config.FRESHNESS_DAYS} calendar days", text)


class RunWindowTests(unittest.TestCase):
    """One window per run, and every layer reads the same one."""

    def tearDown(self):
        from job_agent import profile
        profile.reset()

    def _run(self, **kwargs):
        from dataclasses import replace
        custom = replace(flutter_uk_profile(), key="compiled:test")
        return pipeline.run(offline=True, verify_live=False, use_llm=False,
                            search_profile=custom, **kwargs)

    def test_the_window_asked_for_is_the_window_reported(self):
        self.assertEqual(self._run(days=7).stats.freshness_days, 7)

    def test_the_window_is_pinned_to_the_active_search_for_the_boards(self):
        from job_agent import profile
        self._run(days=7)
        self.assertEqual(profile.active().freshness_days, 7)

    def test_the_default_window_stands_when_the_run_names_none(self):
        self.assertEqual(self._run().stats.freshness_days, config.FRESHNESS_DAYS)


class OnSiteWorkMustBeReachableTests(unittest.TestCase):
    """Work that requires attendance is judged on where the office is.

    `assess_remote` only reads location on the remote path, because a remote
    advert is the only one whose location it can argue with. A search that
    accepts on-site or hybrid work skipped that path entirely, so a Germany
    search returned on-site roles in Singapore, Texas and Melbourne.
    """

    def tearDown(self):
        profile.reset()

    def _searching_from(self, country: str, arrangement: str = "any"):
        from dataclasses import replace
        from job_agent import region
        built = region.build(country)
        profile.set_active(replace(profile.active(), key="test",
                                   work_arrangement=arrangement,
                                   home_country=built.country,
                                   home_terms=built.terms,
                                   home_city_terms=built.cities))

    def _advert(self, location: str, description: str):
        return RawJob(source="manual", source_id="m1",
                      title="Senior Full Stack Developer", company="Acme",
                      url="https://example.com/job", description=description,
                      location_raw=location, posted_at=now_local())

    def test_an_on_site_role_abroad_is_rejected(self):
        from job_agent.pipeline import judge_arrangement
        cases = [
            ("Germany", "Singapore, Singapore", "Five days a week in our office."),
            ("Germany", "Texas (USA)", "On site at our Westlake campus."),
            ("Australia", "Berlin, Deutschland", "Work from our Berlin office."),
            ("Canada", "London", "Two days a week in our London office."),
            ("Saudi Arabia", "Melbourne, Victoria, Australia", "On site, full time."),
            ("Austria", "Toronto, Ontario, Canada", "In our Toronto office daily."),
        ]
        for home, location, description in cases:
            with self.subTest(home=home, location=location):
                self._searching_from(home)
                verdict = judge_arrangement(self._advert(location, description), "any")
                self.assertFalse(verdict.passed, verdict.remote_status)
                self.assertEqual(verdict.category, "ineligible")

    def test_an_on_site_role_at_home_still_passes(self):
        from job_agent.pipeline import judge_arrangement
        cases = [
            ("Germany", "Berlin, Deutschland"),
            ("Germany", "Stuttgart, Baden-Wurttemberg"),
            ("Australia", "Melbourne, Victoria, Australia"),
            ("Canada", "Toronto, Ontario, Canada"),
            ("Saudi Arabia", "Riyadh, Saudi Arabia"),
            ("United Kingdom", "London"),
        ]
        for home, location in cases:
            with self.subTest(home=home, location=location):
                self._searching_from(home)
                verdict = judge_arrangement(
                    self._advert(location, "Five days a week in the office."), "any")
                self.assertTrue(verdict.passed, verdict.reason)

    def test_a_location_that_names_nowhere_is_left_to_the_other_gates(self):
        from job_agent.pipeline import judge_arrangement
        self._searching_from("Germany")
        verdict = judge_arrangement(
            self._advert("Somewhere Nice", "On site, five days a week."), "any")
        self.assertTrue(verdict.passed, verdict.reason)
